"""
__all__ = [
    'add_chart',
    'add_conditional_formatting',
    'add_data',
    'add_formula',
    'auto_filter',
    'create_budget_report',
    'create_data_table',
    'create_project_tracker',
    'create_sales_report',
    'create_sheet',
    'format_range',
    'freeze_panes',
    'protect_sheet',
    'save',
    'set_active_sheet',
    'set_column_widths',
]

Excel 表格generate器 - Excel Generator
基于 openpyxl 的表格generate功能
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from loguru import logger

@dataclass
class ColumnConfig:
    """列配置"""
    header: str
    width: Optional[float] = None
    format: Optional[str] = None
    alignment: str = "left"

@dataclass
class ChartConfig:
    """图表配置"""
    chart_type: str  # 'bar', 'pie', 'line'
    title: str
    data_range: str
    categories_range: Optional[str] = None
    position: Optional[str] = None

class ExcelGenerator:
    """
    Excel 表格generate器
    
    功能:
    - 创建和编辑 Excel 工作簿
    - 数据导入和格式化
    - 图表generate
    - 公式计算
    """
    
    def __init__(self, template_path: Optional[str] = None):
        """
        initgenerate器
        
        Args:
            template_path: 可选的模板文件路径
        """
        if template_path and Path(template_path).exists():
            self.wb = load_workbook(template_path)
            logger.info(f"加载模板: {template_path}")
        else:
            self.wb = Workbook()
            # 移除默认工作表
            if 'Sheet' in self.wb.sheetnames:
                self.wb.remove(self.wb['Sheet'])
        
        self.template_path = template_path
        self.current_sheet = None
    
    def create_sheet(self, title: str, index: Optional[int] = None) -> Any:
        """
        创建工作表
        
        Args:
            title: 工作表标题
            index: 插入位置
        
        Returns:
            工作表对象
        """
        self.current_sheet = self.wb.create_sheet(title=title, index=index)
        return self.current_sheet
    
    def set_active_sheet(self, title: str):
        """设置当前工作表"""
        if title in self.wb.sheetnames:
            self.current_sheet = self.wb[title]
        else:
            raise ValueError(f"工作表不存在: {title}")
    
    def add_data(
        self,
        data: Union[List[List], pd.DataFrame],
        headers: Optional[List[str]] = None,
        start_row: int = 1,
        start_col: int = 1
    ):
        """
        添加数据到工作表
        
        Args:
            data: 数据(列表或 DataFrame)
            headers: 表头
            start_row: 起始行
            start_col: 起始列
        """
        if self.current_sheet is None:
            self.create_sheet("Sheet1")
        
        sheet = self.current_sheet
        
        # 处理 DataFrame
        if isinstance(data, pd.DataFrame):
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, start_col):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        else:
            # 处理列表
            current_row = start_row
            
            # 写入表头
            if headers:
                for col_idx, header in enumerate(headers, start_col):
                    cell = sheet.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                current_row += 1
            
            # 写入数据
            for row_data in data:
                for col_idx, value in enumerate(row_data, start_col):
                    sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
    
    def format_range(
        self,
        cell_range: str,
        font: Optional[Dict] = None,
        alignment: Optional[Dict] = None,
        fill: Optional[Dict] = None,
        border: Optional[Dict] = None
    ):
        """
        格式化单元格范围
        
        Args:
            cell_range: 单元格范围(如 'A1:D10')
            font: 字体设置
            alignment: 对齐设置
            fill: 填充设置
            border: 边框设置
        """
        if self.current_sheet is None:
            return
        
        sheet = self.current_sheet
        
        for row in sheet[cell_range]:
            for cell in row:
                if font:
                    cell.font = Font(**font)
                if alignment:
                    align_map = {
                        'left': 'left',
                        'center': 'center',
                        'right': 'right'
                    }
                    cell.alignment = Alignment(
                        horizontal=align_map.get(alignment.get('horizontal'), 'left'),
                        vertical=alignment.get('vertical', 'center')
                    )
                if fill:
                    cell.fill = PatternFill(**fill)
                if border:
                    cell.border = Border(**border)
    
    def set_column_widths(self, widths: Dict[str, float]):
        """
        设置列宽
        
        Args:
            widths: 列宽字典,如 {'A': 20, 'B': 30}
        """
        if self.current_sheet is None:
            return
        
        for col, width in widths.items():
            self.current_sheet.column_dimensions[col].width = width
    
    def add_formula(self, cell: str, formula: str):
        """
        添加公式
        
        Args:
            cell: 单元格位置
            formula: 公式
        """
        if self.current_sheet is None:
            return
        
        self.current_sheet[cell] = formula
    
    def add_chart(self, config: ChartConfig):
        """
        添加图表
        
        Args:
            config: 图表配置
        """
        if self.current_sheet is None:
            return
        
        sheet = self.current_sheet
        
        # 创建图表
        if config.chart_type == 'bar':
            chart = BarChart()
        elif config.chart_type == 'pie':
            chart = PieChart()
        elif config.chart_type == 'line':
            chart = LineChart()
        else:
            raise ValueError(f"不支持的图表类型: {config.chart_type}")
        
        chart.title = config.title
        
        # 设置数据范围
        data = Reference(sheet, min_col=1, min_row=1, max_row=10, max_col=2)
        chart.add_data(data, titles_from_data=True)
        
        # 设置类别
        if config.categories_range:
            cats = Reference(sheet, min_col=1, min_row=2, max_row=10)
            chart.set_categories(cats)
        
        # 添加图表到工作表
        position = config.position or "E5"
        sheet.add_chart(chart, position)
    
    def add_conditional_formatting(
        self,
        cell_range: str,
        rule_type: str = "color_scale",
        **kwargs
    ):
        """
        添加条件格式
        
        Args:
            cell_range: 单元格范围
            rule_type: 规则类型
        """
        from openpyxl.formatting.rule import ColorScaleRule
        
        if self.current_sheet is None:
            return
        
        if rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type='min',
                start_color='F8696B',
                end_type='max',
                end_color='63BE7B'
            )
            self.current_sheet.conditional_formatting.add(cell_range, rule)
    
    def auto_filter(self, cell_range: str):
        """
        添加自动筛选
        
        Args:
            cell_range: 单元格范围
        """
        if self.current_sheet is None:
            return
        
        self.current_sheet.auto_filter.ref = cell_range
    
    def freeze_panes(self, cell: str):
        """
        冻结窗格
        
        Args:
            cell: 单元格位置
        """
        if self.current_sheet is None:
            return
        
        self.current_sheet.freeze_panes = cell
    
    def protect_sheet(self, password: Optional[str] = None):
        """
        保护工作表
        
        Args:
            password: 保护密码
        """
        if self.current_sheet is None:
            return
        
        self.current_sheet.protection.sheet = True
        if password:
            self.current_sheet.protection.password = password
    
    def create_data_table(
        self,
        title: str,
        headers: List[str],
        data: List[List],
        column_widths: Optional[Dict[str, float]] = None,
        add_filter: bool = True,
        freeze_header: bool = True
    ):
        """
        创建数据表
        
        Args:
            title: 表格标题
            headers: 表头
            data: 数据
            column_widths: 列宽
            add_filter: 是否添加筛选
            freeze_header: 是否冻结表头
        """
        self.create_sheet(title)
        
        # 添加标题
        self.current_sheet['A1'] = title
        self.current_sheet['A1'].font = Font(bold=True, size=14)
        self.current_sheet.merge_cells('A1:' + chr(64 + len(headers)) + '1')
        
        # 添加数据
        self.add_data(data, headers, start_row=3)
        
        # 设置列宽
        if column_widths:
            self.set_column_widths(column_widths)
        
        # 添加筛选
        if add_filter and headers:
            end_col = chr(64 + len(headers))
            end_row = len(data) + 3
            self.auto_filter(f"A3:{end_col}{end_row}")
        
        # 冻结表头
        if freeze_header:
            self.freeze_panes('A4')
    
    def save(self, output_path: str) -> str:
        """
        保存工作簿
        
        Args:
            output_path: 输出路径
        
        Returns:
            保存的文件路径
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.wb.save(output_path)
        logger.info(f"Excel 已保存: {output_path}")
        
        return str(output_path)

# 预设模板
class ExcelTemplates:
    """Excel 表格模板"""
    
    @staticmethod
    def create_project_tracker(output_path: str) -> str:
        """创建项目跟踪表"""
        gen = ExcelGenerator()
        
        headers = ["任务ID", "任务名称", "负责人", "状态", "开始日期", "截止日期", "进度(%)", "备注"]
        sample_data = [
            ["T001", "需求分析", "张三", "进行中", "2024-01-01", "2024-01-15", 80, ""],
            ["T002", "系统设计", "李四", "待开始", "2024-01-16", "2024-01-30", 0, ""],
            ["T003", "编码实现", "王五", "待开始", "2024-02-01", "2024-02-28", 0, ""],
        ]
        
        column_widths = {
            'A': 10, 'B': 20, 'C': 12, 'D': 12,
            'E': 12, 'F': 12, 'G': 10, 'H': 20
        }
        
        gen.create_data_table(
            title="项目任务跟踪表",
            headers=headers,
            data=sample_data,
            column_widths=column_widths
        )
        
        return gen.save(output_path)
    
    @staticmethod
    def create_budget_report(output_path: str) -> str:
        """创建预算报表"""
        gen = ExcelGenerator()
        
        headers = ["项目", "预算", "实际支出", "剩余", "完成率(%)"]
        data = [
            ["人力成本", 100000, 85000, 15000, 85],
            ["设备采购", 50000, 45000, 5000, 90],
            ["运营费用", 30000, 28000, 2000, 93],
            ["其他", 20000, 15000, 5000, 75],
        ]
        
        gen.create_sheet("预算报表")
        gen.add_data(data, headers)
        
        # 添加汇总行
        gen.current_sheet['A6'] = "总计"
        gen.current_sheet['A6'].font = Font(bold=True)
        gen.add_formula('B6', '=SUM(B2:B5)')
        gen.add_formula('C6', '=SUM(C2:C5)')
        gen.add_formula('D6', '=SUM(D2:D5)')
        gen.add_formula('E6', '=AVERAGE(E2:E5)')
        
        # 设置列宽
        gen.set_column_widths({'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 12})
        
        return gen.save(output_path)
    
    @staticmethod
    def create_sales_report(output_path: str) -> str:
        """创建销售报表"""
        gen = ExcelGenerator()
        
        # 销售数据
        headers = ["月份", "产品A", "产品B", "产品C", "总计"]
        data = [
            ["1月", 120, 150, 80, "=SUM(B2:D2)"],
            ["2月", 135, 165, 90, "=SUM(B3:D3)"],
            ["3月", 150, 180, 100, "=SUM(B4:D4)"],
            ["4月", 145, 175, 95, "=SUM(B5:D5)"],
        ]
        
        gen.create_sheet("销售报表")
        gen.add_data(data, headers)
        
        # 添加图表
        chart_config = ChartConfig(
            chart_type='bar',
            title='月度销售趋势',
            data_range='A1:D5',
            categories_range='A2:A5',
            position='F2'
        )
        gen.add_chart(chart_config)
        
        return gen.save(output_path)
